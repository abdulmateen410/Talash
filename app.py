import os
import re
import json
import time
import uuid
import base64
import hashlib
import threading
import requests
from datetime import datetime
from pathlib import Path
from functools import wraps
from collections import Counter, defaultdict
from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for

from dotenv import load_dotenv
load_dotenv()

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from groq import Groq
    GROQ_OK = True
except ImportError:
    GROQ_OK = False

try:
    from PIL import Image
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import pytesseract
    TESSERACT_OK = True
except ImportError:
    TESSERACT_OK = False

try:
    import fitz  # PyMuPDF
    FITZ_OK = True
except ImportError:
    FITZ_OK = False

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.cluster import KMeans
    SKLEARN_OK = True
except ImportError:
    SKLEARN_OK = False

try:
    from pymongo import MongoClient
    from bson import ObjectId
    MONGO_OK = True
except ImportError:
    MONGO_OK = False

import numpy as np

class NumpyJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder that handles numpy types globally."""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, np.bool_):
            return bool(obj)
        return super().default(obj)

app = Flask(__name__)
app.json_encoder = NumpyJSONEncoder
app.secret_key = os.environ.get("SECRET_KEY", "talash-secret-key-2026")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024

BASE_DIR   = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
USERS_FILE = BASE_DIR / "users.json"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# MongoDB Connection
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017/")
MONGO_DB_NAME = "Talash_db"

mongo_client = None
db = None

if MONGO_OK:
    try:
        mongo_client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
        mongo_client.server_info()  # Test connection
        db = mongo_client[MONGO_DB_NAME]
        print(f"✓ MongoDB connected: {MONGO_DB_NAME}")
    except Exception as e:
        print(f"✗ MongoDB connection failed: {e}")
        MONGO_OK = False

jobs      = {}
jobs_lock = threading.Lock()

ALLOWED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp"}

# ─── Auth helpers ─────────────────────────────────────────────────────────────

def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def load_users() -> dict:
    if USERS_FILE.exists():
        try:
            return json.loads(USERS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_users(users: dict):
    USERS_FILE.write_text(json.dumps(users, indent=2, ensure_ascii=False), encoding="utf-8")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "username" not in session:
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Authentication required"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

# ─── MongoDB helpers ──────────────────────────────────────────────────────────

def convert_numpy_types(obj):
    """Convert numpy types to Python native types for MongoDB compatibility."""
    if isinstance(obj, dict):
        return {str(k): convert_numpy_types(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return convert_numpy_types(obj.tolist())
    elif isinstance(obj, np.bool_):
        return bool(obj)
    else:
        return obj


def normalize_grade_to_percentage(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace(",", ".")
    # Handle fractional formats like 3.2/4.0 or 8.5/10
    if "/" in text:
        parts = [p.strip() for p in text.split("/") if p.strip()]
        if len(parts) >= 2:
            try:
                num = float(re.findall(r"\d+\.?\d*", parts[0])[0])
                den = float(re.findall(r"\d+\.?\d*", parts[1])[0])
                if den > 0:
                    return f"{round(num / den * 100, 1)}%"
            except Exception:
                pass
    # Handle explicit percentages
    if "%" in text:
        try:
            num = float(re.findall(r"\d+\.?\d*", text)[0])
            return f"{round(num, 1)}%"
        except Exception:
            return text
    # Parse plain numeric values
    try:
        num = float(re.findall(r"\d+\.?\d*", text)[0])
    except Exception:
        return text
    if num <= 4.0:
        return f"{round(num * 25, 1)}%"
    if num <= 10.0:
        return f"{round(num * 10, 1)}%"
    return f"{round(num, 1)}%"


def normalize_degree_label(level, degree):
    raw = f"{level or ''} {degree or ''}".strip().lower()
    if not raw:
        return "Unknown"
    if "phd" in raw or "doctor" in raw or "doctorate" in raw:
        return "PhD"
    if "master" in raw or "ms" in raw or "m.s" in raw or "m.sc" in raw or "mphil" in raw or "ma" in raw:
        return "MS"
    if "bachelor" in raw or "bs" in raw or "b.sc" in raw or "bsc" in raw:
        return "BS"
    if "fsc" in raw or "hssc" in raw or "a-level" in raw or "alevel" in raw:
        return "FSC/A-Level"
    if "ssc" in raw or "o-level" in raw or "olevel" in raw:
        return "SSC/O-Level"
    if "high school" in raw or "secondary" in raw:
        return "SSC/O-Level"
    if "intermediate" in raw or "higher secondary" in raw:
        return "FSC/A-Level"
    return degree or level or "Unknown"


def normalize_education_entry(entry):
    if not isinstance(entry, dict):
        return entry

    degree = str(entry.get("degree", "")).strip()
    level = str(entry.get("level", "")).strip()
    normalized_degree = normalize_degree_label(level, degree)
    normalized_level = normalized_degree if normalized_degree in {"PhD", "MS", "BS", "FSC/A-Level", "SSC/O-Level"} else level or normalized_degree
    normalized_grade = normalize_grade_to_percentage(entry.get("grade_cgpa_percentage", ""))

    return {
        **entry,
        "degree": normalized_degree,
        "level": normalized_level,
        "grade_cgpa_percentage": normalized_grade
    }


def get_education_weight(level, degree):
    label = normalize_degree_label(level, degree).lower()
    order = {
        "phd": 5,
        "ms": 4,
        "bs": 3,
        "fsc/a-level": 2,
        "ssc/o-level": 1,
        "unknown": 0
    }
    return order.get(label, 0)


def save_candidate_to_db(username, candidate_data):
    """Save candidate to MongoDB"""
    if not MONGO_OK or db is None:
        return None
    
    try:
        # Convert any numpy types to Python native types
        candidate_data = convert_numpy_types(candidate_data)

        if isinstance(candidate_data.get("education"), list):
            candidate_data["education"] = [
                normalize_education_entry(entry)
                for entry in candidate_data["education"]
            ]
        
        candidate_data["username"] = username
        candidate_data["created_at"] = datetime.now()
        candidate_data["updated_at"] = datetime.now()
        
        result = db.candidates.insert_one(candidate_data)
        return str(result.inserted_id)
    except Exception as e:
        print(f"Error saving to MongoDB: {e}")
        return None

def get_all_candidates(username):
    """Get all candidates for a user from MongoDB"""
    if not MONGO_OK or db is None:
        return []
    
    try:
        candidates = list(db.candidates.find({"username": username}).sort("created_at", -1))
        
        # Convert ObjectId to string
        for c in candidates:
            c["_id"] = str(c["_id"])
        
        return candidates
    except Exception as e:
        print(f"Error fetching from MongoDB: {e}")
        return []

def delete_candidate(username, candidate_id):
    """Delete a candidate from MongoDB"""
    if not MONGO_OK or db is None:
        return False
    
    try:
        result = db.candidates.delete_one({
            "_id": ObjectId(candidate_id),
            "username": username
        })
        return result.deleted_count > 0
    except Exception as e:
        print(f"Error deleting from MongoDB: {e}")
        return False

# ─── Extraction prompts ───────────────────────────────────────────────────────

EXTRACTION_PROMPT = """\
You are an expert HR CV data extraction system.

Extract ALL structured information from the CV text below.
Return ONLY valid JSON — no markdown fences, no explanation, no extra keys.

CV TEXT:
{cv_text}

Return EXACTLY this JSON structure:
{{
  "personal_info": {{
    "name":"", "father_name":"", "dob":"", "nationality":"",
    "marital_status":"", "cnic":"", "email":"", "phone":"",
    "address":"", "current_salary":"", "expected_salary":"",
    "present_employment":"", "apply_date":"", "post_applied":""
  }},
  "education": [{{
    "degree":"", "level":"", "specialization":"",
    "grade_cgpa_percentage":"", "passing_year":"", "start_year":"",
    "institution":"", "board_university":"", "location":""
  }}],
  "professional_qualifications": [{{"qualification":"", "passing_year":"", "institution":""}}],
  "experience": [{{
    "post":"", "organization":"", "location":"",
    "start_date":"", "end_date":"", "duration_months":null, 
    "employment_type":"", "responsibilities":""
  }}],
  "publications": [{{
    "title":"", "authors":"", "author_position":"",
    "published_in":"", "venue_type":"", "issn":"", "isbn":"",
    "impact_factor":null, "volume":"", "issue":"", "pages":"", 
    "year":"", "type":"", "doi":"", "url":""
  }}],
  "supervision": {{
    "ms_main": [], "ms_co": [], "phd_main": [], "phd_co": [],
    "publications_with_students": []
  }},
  "patents": [{{
    "patent_number":"", "title":"", "date":"", 
    "inventors":"", "country":"", "link":"", "status":""
  }}],
  "books": [{{
    "title":"", "authors":"", "isbn":"", "publisher":"", 
    "year":"", "link":"", "role":""
  }}],
  "skills": {{
    "technical": [], "research": [], "teaching": [], 
    "languages": [], "certifications": []
  }},
  "awards_scholarships": [{{"type":"", "detail":"", "year":""}}],
  "references": [{{
    "name":"", "designation":"", "organization":"", 
    "email":"", "phone":"", "relationship":""
  }}]
}}

Rules:
- level must be one of: SSC | HSSC | Bachelor | Master | MPhil | MS | PhD | PostDoc | Other
- type (publications) must be: Journal | Conference | Book Chapter | Workshop | Other
- venue_type must be: Journal | Conference | Workshop | Other
- Supervision arrays should contain student names and graduation years
- All missing fields = "" or null or []; no extra keys
- Extract ALL entries, do not truncate lists
"""

IMAGE_PROMPT = """\
This is a CV / resume image. Extract ALL text visible in the image exactly as it appears,
preserving structure and layout as much as possible. Include all sections:
personal info, education, experience, publications, supervision, patents, books, 
awards, references, etc.
Output only the raw extracted text — no commentary.
"""

ANALYSIS_PROMPT = """\
Analyze this candidate's complete profile and generate a comprehensive assessment.

PROFILE DATA:
{profile_json}

Provide a structured analysis covering:
1. Educational strength and quality (performance, institutions, progression, gaps)
2. Research output quality (journal rankings, conference quality, authorship roles)
3. Research breadth vs depth (topic variability)
4. Collaboration patterns (co-authorship analysis)
5. Professional experience consistency and progression
6. Skill alignment with claimed expertise
7. Overall suitability assessment
8. Key strengths and concerns

Return as valid JSON:
{{
  "overall_score": 0-100,
  "education_score": 0-100,
  "research_score": 0-100,
  "experience_score": 0-100,
  "summary": "2-3 paragraph comprehensive summary",
  "strengths": ["strength 1", "strength 2", ...],
  "concerns": ["concern 1", "concern 2", ...],
  "recommendations": "hiring recommendation",
  "missing_info": ["item 1", "item 2", ...]
}}
"""

# ─── Text extraction ──────────────────────────────────────────────────────────

def extract_text_from_pdf(pdf_path):
    """Try pdfplumber for text-layer PDFs."""
    if not PDF_OK:
        return ""
    pages = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text and text.strip():
                    pages.append(f"[PAGE {i+1}]\n{text}")
    except Exception:
        pass
    return "\n\n".join(pages)

def pdf_to_pil_images(pdf_path, dpi=250):
    """Render PDF pages to PIL Images using PyMuPDF."""
    if not FITZ_OK or not PIL_OK:
        return []
    images = []
    try:
        doc = fitz.open(pdf_path)
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        for page in doc:
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
        doc.close()
    except Exception:
        pass
    return images

def pil_image_to_base64(img, fmt="PNG"):
    """Convert PIL Image to base64 string."""
    import io
    buf = io.BytesIO()
    img.save(buf, format=fmt)
    return base64.b64encode(buf.getvalue()).decode("utf-8")

def ocr_image_tesseract(img):
    """Run Tesseract OCR on a PIL Image."""
    if not TESSERACT_OK:
        return ""
    try:
        custom_cfg = r"--oem 3 --psm 1"
        return pytesseract.image_to_string(img, lang="eng", config=custom_cfg).strip()
    except Exception:
        return ""

def ocr_image_groq_vision(client, img):
    """Use Groq vision model to extract text from image."""
    if not client:
        return ""
    try:
        b64 = pil_image_to_base64(img, fmt="PNG")
        response = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": IMAGE_PROMPT},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}}
                ]
            }],
            max_tokens=4096,
            temperature=0,
        )
        return response.choices[0].message.content.strip()
    except Exception:
        return ""

def extract_text_from_scanned_pdf(pdf_path, client=None):
    """Extract text from scanned PDF using OCR."""
    images = pdf_to_pil_images(pdf_path, dpi=250)
    if not images:
        return ""
    
    all_text = []
    for i, img in enumerate(images):
        if client:
            text = ocr_image_groq_vision(client, img)
        else:
            text = ocr_image_tesseract(img)
        
        if text.strip():
            all_text.append(f"[PAGE {i+1}]\n{text}")
    
    return "\n\n".join(all_text)

def extract_text_from_image(img_path, client=None):
    """Extract text from an image file."""
    if not PIL_OK:
        return ""
    try:
        img = Image.open(img_path).convert("RGB")
        if client:
            return ocr_image_groq_vision(client, img)
        else:
            return ocr_image_tesseract(img)
    except Exception:
        return ""

def extract_text_from_file(file_path, client=None):
    """Main dispatcher for text extraction."""
    ext = Path(file_path).suffix.lower()
    
    if ext == ".pdf":
        text = extract_text_from_pdf(file_path)
        if len(text.strip()) < 100:
            text = extract_text_from_scanned_pdf(file_path, client)
        return text
    elif ext in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp"}:
        return extract_text_from_image(file_path, client)
    
    return ""

# ─── LLM helpers ──────────────────────────────────────────────────────────────

def init_groq_client(api_key):
    """Initialize Groq client if API key is available."""
    if not GROQ_OK or not api_key:
        return None
    try:
        return Groq(api_key=api_key)
    except Exception:
        return None

def call_groq(client, prompt, max_tokens=8000):
    """Call Groq LLM with prompt."""
    if not client:
        return ""
    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=max_tokens,
            temperature=0,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"Error: {e}"

def parse_json_response(text):
    """Extract and parse JSON from LLM response."""
    text = text.strip()
    
    # Remove markdown code fences
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    
    # Try to find JSON object
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        text = match.group(0)
    
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None

# ─── Research Profile Analysis (Same as before) ───────────────────────────────

def analyze_journal_publication(pub, client=None):
    analysis = {
        "title": pub.get("title", ""),
        "journal": pub.get("published_in", ""),
        "year": pub.get("year", ""),
        "wos_indexed": False,
        "scopus_indexed": False,
        "impact_factor": pub.get("impact_factor"),
        "quartile": "",
        "legitimacy": "Unknown",
        "author_role": determine_author_role(pub),
        "quality_score": 0
    }
    
    if client:
        prompt = f"""
Analyze this journal publication and estimate its quality:
Title: {pub.get('title', '')}
Journal: {pub.get('published_in', '')}
Year: {pub.get('year', '')}
ISSN: {pub.get('issn', '')}

Provide quality estimation as JSON:
{{
  "wos_indexed": true/false,
  "scopus_indexed": true/false,
  "quartile": "Q1"|"Q2"|"Q3"|"Q4"|"Unknown",
  "legitimacy": "High"|"Medium"|"Low"|"Unknown",
  "quality_score": 0-100
}}
"""
        response = call_groq(client, prompt, max_tokens=500)
        result = parse_json_response(response)
        if result:
            analysis.update(result)
    
    return analysis

def analyze_conference_publication(pub, client=None):
    analysis = {
        "title": pub.get("title", ""),
        "conference": pub.get("published_in", ""),
        "year": pub.get("year", ""),
        "core_rank": "",
        "indexed": False,
        "maturity": "",
        "author_role": determine_author_role(pub),
        "quality_score": 0
    }
    
    if client:
        prompt = f"""
Analyze this conference publication:
Title: {pub.get('title', '')}
Conference: {pub.get('published_in', '')}
Year: {pub.get('year', '')}

Provide quality estimation as JSON:
{{
  "core_rank": "A*"|"A"|"B"|"C"|"Unknown",
  "indexed": true/false,
  "maturity": "Mature (10+ years)"|"Established (5-10 years)"|"New (<5 years)"|"Unknown",
  "quality_score": 0-100
}}
"""
        response = call_groq(client, prompt, max_tokens=500)
        result = parse_json_response(response)
        if result:
            analysis.update(result)
    
    return analysis

def determine_author_role(pub):
    authors = pub.get("authors", "")
    position = pub.get("author_position", "")
    
    if not authors:
        return "Unknown"
    
    author_list = [a.strip() for a in authors.split(",")]
    
    if position:
        pos_lower = position.lower()
        if "first" in pos_lower:
            return "First Author"
        elif "corresponding" in pos_lower:
            return "Corresponding Author"
        elif "last" in pos_lower:
            return "Last Author"
    
    if len(author_list) == 1:
        return "Sole Author"
    
    return "Co-Author"

def analyze_topic_variability(publications, client=None):
    if not publications or not SKLEARN_OK:
        return {
            "topics": [],
            "diversity_score": 0,
            "focus": "Unknown"
        }
    
    texts = []
    for pub in publications:
        title = pub.get("title", "")
        texts.append(title)
    
    if len(texts) < 3:
        return {
            "topics": ["Insufficient data"],
            "diversity_score": 0,
            "focus": "Limited publications"
        }
    
    try:
        vectorizer = TfidfVectorizer(max_features=50, stop_words='english')
        X = vectorizer.fit_transform(texts)
        
        n_clusters = min(5, len(texts) // 2 + 1)
        kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
        labels = kmeans.fit_predict(X)
        
        cluster_counts = Counter(int(k) for k in labels)
        
        total = len(publications)
        diversity = (len(cluster_counts) / total) * 100 if total > 0 else 0
        
        max_cluster_size = max(cluster_counts.values())
        focus_ratio = max_cluster_size / total if total > 0 else 0
        
        if focus_ratio > 0.7:
            focus = "Highly Focused"
        elif focus_ratio > 0.5:
            focus = "Moderately Focused"
        else:
            focus = "Diverse/Interdisciplinary"
        
        topics = []
        feature_names = vectorizer.get_feature_names_out()
        for i in range(n_clusters):
            center = kmeans.cluster_centers_[i]
            top_indices = center.argsort()[-3:][::-1]
            top_terms = [feature_names[idx] for idx in top_indices]
            topics.append(" | ".join(top_terms))
        
        return {
            "topics": topics,
            "diversity_score": round(diversity, 1),
            "focus": focus,
            "cluster_distribution": dict(cluster_counts)
        }
    except Exception:
        return {
            "topics": ["Analysis failed"],
            "diversity_score": 0,
            "focus": "Unknown"
        }

def analyze_coauthorship(publications):
    if not publications:
        return {
            "total_coauthors": 0,
            "frequent_collaborators": [],
            "avg_coauthors_per_paper": 0,
            "collaboration_score": 0
        }
    
    all_coauthors = []
    papers_with_coauthors = 0
    
    for pub in publications:
        authors = pub.get("authors", "")
        if not authors:
            continue
        
        author_list = [a.strip() for a in authors.split(",")]
        if len(author_list) > 1:
            papers_with_coauthors += 1
            all_coauthors.extend(author_list)
    
    coauthor_counts = Counter(all_coauthors)
    
    if coauthor_counts:
        most_common_name = coauthor_counts.most_common(1)[0][0]
        if coauthor_counts[most_common_name] > len(publications) * 0.8:
            del coauthor_counts[most_common_name]
    
    top_collaborators = coauthor_counts.most_common(10)
    
    unique_coauthors = len(coauthor_counts)
    avg_coauthors = len(all_coauthors) / len(publications) if publications else 0
    collaboration_score = min(100, (unique_coauthors / len(publications)) * 50) if publications else 0
    
    return {
        "total_coauthors": unique_coauthors,
        "frequent_collaborators": [{"name": name, "count": count} for name, count in top_collaborators[:5]],
        "avg_coauthors_per_paper": round(avg_coauthors, 1),
        "collaboration_score": round(collaboration_score, 1),
        "collaboration_breadth": "High" if unique_coauthors > len(publications) else "Moderate" if unique_coauthors > len(publications) * 0.5 else "Low"
    }

# ─── Educational & Experience Analysis (Same as before) ───────────────────────

def analyze_educational_profile(education, client=None):
    if not education:
        return {
            "highest_degree": "None",
            "performance_trend": "Unknown",
            "institutional_quality": [],
            "gaps": [],
            "progression_score": 0
        }
    
    normalized_education = []
    for edu in education:
        degree = str(edu.get("degree", "")).strip()
        level = str(edu.get("level", "")).strip()
        normalized_degree = normalize_degree_label(level, degree)
        normalized_level = normalized_degree if normalized_degree in {"PhD", "MS", "BS", "FSC/A-Level", "SSC/O-Level"} else level or normalized_degree
        normalized_grade = normalize_grade_to_percentage(edu.get("grade_cgpa_percentage", ""))

        normalized_education.append({
            **edu,
            "degree": normalized_degree,
            "level": normalized_level,
            "grade_cgpa_percentage": normalized_grade
        })
    
    sorted_edu = sorted(
        normalized_education,
        key=lambda x: get_education_weight(x.get("level", ""), x.get("degree", ""))
    )
    
    highest = sorted_edu[-1] if sorted_edu else {}
    highest_degree = highest.get("degree", "Unknown")
    
    grades = []
    for edu in sorted_edu:
        grade_str = edu.get("grade_cgpa_percentage", "")
        if grade_str:
            try:
                nums = re.findall(r"\d+\.?\d*", grade_str)
                if nums:
                    val = float(nums[0])
                    if grade_str.strip().endswith("%"):
                        grades.append(val)
                    else:
                        grades.append(val)
            except Exception:
                pass
    
    performance_trend = "Stable"
    if len(grades) >= 2:
        if grades[-1] > grades[0] * 1.1:
            performance_trend = "Improving"
        elif grades[-1] < grades[0] * 0.9:
            performance_trend = "Declining"
    
    institutional_quality = []
    for edu in sorted_edu:
        inst = edu.get("institution", "")
        if inst:
            institutional_quality.append({
                "institution": inst,
                "ranking": "Not verified",
                "level": edu.get("level", "")
            })
    
    gaps = detect_educational_gaps(sorted_edu)
    progression_score = calculate_progression_score(sorted_edu, gaps)
    
    return {
        "highest_degree": highest_degree,
        "performance_trend": performance_trend,
        "institutional_quality": institutional_quality,
        "gaps": gaps,
        "progression_score": progression_score
    }


def detect_educational_gaps(education):
    gaps = []
    
    for i in range(len(education) - 1):
        current = education[i]
        next_edu = education[i + 1]
        
        current_end = current.get("passing_year", "")
        next_start = next_edu.get("start_year", "") or next_edu.get("passing_year", "")
        
        if current_end and next_start:
            try:
                end_year = int(current_end)
                start_year = int(next_start)
                gap_years = start_year - end_year - 1
                
                if gap_years > 0:
                    gaps.append({
                        "after": current.get("level", ""),
                        "before": next_edu.get("level", ""),
                        "duration_years": gap_years,
                        "justified": False
                    })
            except:
                pass
    
    return gaps

def calculate_progression_score(education, gaps):
    score = 50
    
    level_bonus = {"Bachelor": 10, "Master": 20, "MPhil": 25, "MS": 25, "PhD": 30, "PostDoc": 35}
    if education:
        highest_level = education[-1].get("level", "")
        score += level_bonus.get(highest_level, 0)
    
    total_gap_years = sum(g.get("duration_years", 0) for g in gaps if not g.get("justified", False))
    score -= min(20, total_gap_years * 5)
    
    if len(education) >= 3 and not gaps:
        score += 10
    
    return max(0, min(100, score))

def analyze_professional_experience(experience, education):
    if not experience:
        return {
            "total_years": 0,
            "overlaps": [],
            "gaps": [],
            "progression": "Unknown",
            "consistency_score": 0
        }
    
    parsed_exp = []
    for exp in experience:
        start = parse_year(exp.get("start_date", ""))
        end = parse_year(exp.get("end_date", "")) or datetime.now().year
        
        if start:
            parsed_exp.append({
                "post": exp.get("post", ""),
                "org": exp.get("organization", ""),
                "start": start,
                "end": end,
                "original": exp
            })
    
    parsed_exp.sort(key=lambda x: x["start"])
    
    overlaps = detect_employment_overlaps(parsed_exp)
    gaps = detect_employment_gaps(parsed_exp, education)
    total_years = calculate_total_experience_years(parsed_exp)
    progression = analyze_career_progression(parsed_exp)
    consistency_score = calculate_experience_consistency(parsed_exp, overlaps, gaps)
    
    return {
        "total_years": total_years,
        "overlaps": overlaps,
        "gaps": gaps,
        "progression": progression,
        "consistency_score": consistency_score
    }

def parse_year(date_str):
    if not date_str:
        return None
    
    patterns = [
        r'(\d{4})',
        r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, date_str)
        if match:
            return int(match.group(1) if len(match.groups()) == 1 else match.group(3))
    
    return None

def detect_employment_overlaps(parsed_exp):
    overlaps = []
    
    for i in range(len(parsed_exp)):
        for j in range(i + 1, len(parsed_exp)):
            exp1 = parsed_exp[i]
            exp2 = parsed_exp[j]
            
            if exp1["start"] <= exp2["end"] and exp2["start"] <= exp1["end"]:
                overlaps.append({
                    "job1": exp1["post"],
                    "job2": exp2["post"],
                    "period": f"{max(exp1['start'], exp2['start'])}-{min(exp1['end'], exp2['end'])}"
                })
    
    return overlaps

def detect_employment_gaps(parsed_exp, education):
    gaps = []
    
    edu_end_year = 0
    if education:
        for edu in education:
            year = parse_year(edu.get("passing_year", ""))
            if year:
                edu_end_year = max(edu_end_year, year)
    
    if parsed_exp and edu_end_year:
        first_job_year = parsed_exp[0]["start"]
        gap_years = first_job_year - edu_end_year
        
        if gap_years > 1:
            gaps.append({
                "type": "Post-education gap",
                "duration_years": gap_years,
                "period": f"{edu_end_year}-{first_job_year}"
            })
    
    for i in range(len(parsed_exp) - 1):
        current_end = parsed_exp[i]["end"]
        next_start = parsed_exp[i + 1]["start"]
        gap_years = next_start - current_end
        
        if gap_years > 1:
            gaps.append({
                "type": "Between jobs",
                "duration_years": gap_years,
                "period": f"{current_end}-{next_start}"
            })
    
    return gaps

def calculate_total_experience_years(parsed_exp):
    if not parsed_exp:
        return 0
    
    periods = [(exp["start"], exp["end"]) for exp in parsed_exp]
    periods.sort()
    
    merged = []
    for start, end in periods:
        if merged and start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
        else:
            merged.append((start, end))
    
    total = sum(end - start for start, end in merged)
    return round(total, 1)

def analyze_career_progression(parsed_exp):
    if len(parsed_exp) < 2:
        return "Insufficient data"
    
    seniority_keywords = {
        "junior": 1, "assistant": 2, "associate": 3, "senior": 4,
        "lead": 5, "principal": 6, "manager": 7, "director": 8, "vp": 9, "chief": 10
    }
    
    scores = []
    for exp in parsed_exp:
        post = exp["post"].lower()
        score = 0
        for keyword, value in seniority_keywords.items():
            if keyword in post:
                score = max(score, value)
        scores.append(score)
    
    if scores[-1] > scores[0]:
        return "Progressive"
    elif scores[-1] < scores[0]:
        return "Regressive"
    else:
        return "Lateral"

def calculate_experience_consistency(parsed_exp, overlaps, gaps):
    score = 80
    
    score -= len(overlaps) * 10
    
    total_gap_years = sum(g.get("duration_years", 0) for g in gaps)
    score -= min(30, total_gap_years * 5)
    
    if len(parsed_exp) >= 3 and not gaps:
        score += 10
    
    return max(0, min(100, score))

def analyze_skill_alignment(candidate_data, client=None):
    skills = candidate_data.get("skills", {})
    experience = candidate_data.get("experience", [])
    publications = candidate_data.get("publications", [])
    
    all_skills = []
    for category in skills.values():
        if isinstance(category, list):
            all_skills.extend(category)
    
    if not all_skills:
        return {
            "evidenced_skills": [],
            "weak_evidence": [],
            "alignment_score": 0
        }
    
    evidenced = []
    weak = []
    
    for skill in all_skills:
        evidence_count = 0
        
        for exp in experience:
            post = exp.get("post", "").lower()
            resp = exp.get("responsibilities", "").lower()
            if skill.lower() in post or skill.lower() in resp:
                evidence_count += 1
        
        for pub in publications:
            title = pub.get("title", "").lower()
            if skill.lower() in title:
                evidence_count += 1
        
        if evidence_count >= 2:
            evidenced.append(skill)
        elif evidence_count == 0:
            weak.append(skill)
    
    alignment_score = (len(evidenced) / len(all_skills) * 100) if all_skills else 0
    
    return {
        "evidenced_skills": evidenced[:10],
        "weak_evidence": weak[:10],
        "alignment_score": round(alignment_score, 1)
    }

# ─── Candidate Ranking System ─────────────────────────────────────────────────

def calculate_candidate_ranking(candidate_data, analyses):
    weights = {
        "education": 0.20,
        "research": 0.35,
        "experience": 0.25,
        "skills": 0.10,
        "collaboration": 0.10
    }
    
    edu_score = analyses.get("educational_analysis", {}).get("progression_score", 0)
    
    pubs = candidate_data.get("publications", [])
    research_score = min(100, len(pubs) * 10)
    
    exp_score = analyses.get("experience_analysis", {}).get("consistency_score", 0)
    skill_score = analyses.get("skill_alignment", {}).get("alignment_score", 0)
    collab_score = analyses.get("coauthorship_analysis", {}).get("collaboration_score", 0)
    
    total_score = (
        edu_score * weights["education"] +
        research_score * weights["research"] +
        exp_score * weights["experience"] +
        skill_score * weights["skills"] +
        collab_score * weights["collaboration"]
    )
    
    return {
        "total_score": round(total_score, 1),
        "education_score": round(edu_score, 1),
        "research_score": round(research_score, 1),
        "experience_score": round(exp_score, 1),
        "skills_score": round(skill_score, 1),
        "collaboration_score": round(collab_score, 1),
        "ranking_tier": get_ranking_tier(total_score)
    }

def get_ranking_tier(score):
    if score >= 80:
        return "Excellent"
    elif score >= 65:
        return "Very Good"
    elif score >= 50:
        return "Good"
    elif score >= 35:
        return "Fair"
    else:
        return "Needs Improvement"

# ─── Comprehensive Candidate Processing ──────────────────────────────────────

def process_candidate_full(client, cv_text):
    if not client:
        return None
    
    prompt = EXTRACTION_PROMPT.format(cv_text=cv_text[:15000])
    response = call_groq(client, prompt, max_tokens=8000)
    candidate_data = parse_json_response(response)
    
    if not candidate_data:
        return None
    
    publications = candidate_data.get("publications", [])
    education = candidate_data.get("education", [])
    experience = candidate_data.get("experience", [])
    
    journal_analyses = []
    conference_analyses = []
    
    for pub in publications:
        pub_type = pub.get("type", "").lower()
        if "journal" in pub_type:
            journal_analyses.append(analyze_journal_publication(pub, client))
        elif "conference" in pub_type:
            conference_analyses.append(analyze_conference_publication(pub, client))
    
    topic_analysis = analyze_topic_variability(publications, client)
    coauthor_analysis = analyze_coauthorship(publications)
    edu_analysis = analyze_educational_profile(education, client)
    exp_analysis = analyze_professional_experience(experience, education)
    skill_analysis = analyze_skill_alignment(candidate_data, client)
    
    analyses = {
        "journal_analyses": journal_analyses,
        "conference_analyses": conference_analyses,
        "topic_variability": topic_analysis,
        "coauthorship_analysis": coauthor_analysis,
        "educational_analysis": edu_analysis,
        "experience_analysis": exp_analysis,
        "skill_alignment": skill_analysis
    }
    
    ranking = calculate_candidate_ranking(candidate_data, analyses)
    summary = generate_candidate_summary(candidate_data, analyses, ranking, client)
    
    result = {
        **candidate_data,
        "analyses": analyses,
        "ranking": ranking,
        "summary": summary
    }
    
    return result

def generate_candidate_summary(candidate_data, analyses, ranking, client):
    """Generate candidate summary — uses Groq if available, otherwise builds from data."""

    name            = (candidate_data.get("personal_info") or {}).get("name", "the candidate")
    highest_degree  = analyses.get("educational_analysis", {}).get("highest_degree", "Unknown")
    total_pubs      = len(candidate_data.get("publications", []))
    exp_years       = analyses.get("experience_analysis", {}).get("total_years", 0)
    focus           = analyses.get("topic_variability", {}).get("focus", "Unknown")
    collab_score    = analyses.get("coauthorship_analysis", {}).get("collaboration_score", 0)
    edu_score       = ranking.get("education_score", 0)
    res_score       = ranking.get("research_score", 0)
    exp_score       = ranking.get("experience_score", 0)
    skill_score     = ranking.get("skills_score", 0)
    total_score     = ranking.get("total_score", 0)
    tier            = ranking.get("ranking_tier", "Unknown")
    progression     = analyses.get("experience_analysis", {}).get("progression", "Unknown")
    evidenced_skills = analyses.get("skill_alignment", {}).get("evidenced_skills", [])
    weak_skills      = analyses.get("skill_alignment", {}).get("weak_evidence", [])
    gaps             = analyses.get("experience_analysis", {}).get("gaps", [])
    edu_gaps         = analyses.get("educational_analysis", {}).get("gaps", [])

    # ── Try Groq first ──────────────────────────────────────────────────────
    if client:
        prompt = f"""
Analyze this academic/professional candidate and generate a structured assessment.

Name: {name}
Highest Degree: {highest_degree}
Total Publications: {total_pubs}
Experience (years): {exp_years}
Research Focus: {focus}
Ranking Score: {total_score} ({tier})

Scores — Education: {edu_score}, Research: {res_score}, Experience: {exp_score}, Skills: {skill_score}, Collaboration: {collab_score}

Return ONLY valid JSON (no markdown):
{{
  "assessment": "2-3 paragraph overall assessment",
  "strengths": ["strength 1", "strength 2", "strength 3"],
  "concerns": ["concern 1", "concern 2", "concern 3"],
  "recommendation": "Strong Hire | Hire | Maybe | No Hire"
}}
"""
        response = call_groq(client, prompt, max_tokens=1500)
        result   = parse_json_response(response)

        if result and result.get("assessment") and result.get("strengths"):
            return result

    # ── Fallback: build summary from data ───────────────────────────────────
    strengths = []
    concerns  = []

    # Education
    if edu_score >= 70:
        strengths.append(f"Strong educational background — holds a {highest_degree}")
    elif edu_score >= 40:
        strengths.append(f"Solid educational foundation with a {highest_degree}")
    else:
        concerns.append(f"Educational profile needs strengthening (score: {edu_score:.0f}/100)")

    # Research
    if total_pubs > 0:
        if res_score >= 70:
            strengths.append(f"Strong research output with {total_pubs} publication(s)")
        else:
            strengths.append(f"Has research experience with {total_pubs} publication(s)")
    else:
        concerns.append("No publications found in the CV")

    # Research focus
    if focus and focus not in ("Unknown", "Insufficient data", "Limited publications"):
        strengths.append(f"Research focus is {focus.lower()}")

    # Experience
    if exp_years > 0:
        if exp_score >= 70:
            strengths.append(f"Extensive professional experience of {exp_years:.0f} year(s) with {progression.lower()} progression")
        else:
            strengths.append(f"Has {exp_years:.0f} year(s) of professional experience")
    else:
        concerns.append("Limited or no professional experience detected")

    # Collaboration
    if collab_score >= 60:
        strengths.append(f"Good collaboration record (score: {collab_score:.0f}/100)")
    elif collab_score < 30 and total_pubs > 0:
        concerns.append("Limited collaboration with other researchers")

    # Skills
    if evidenced_skills:
        strengths.append(f"Verified skills: {', '.join(evidenced_skills[:4])}")
    if weak_skills:
        concerns.append(f"Skills lacking evidence: {', '.join(weak_skills[:3])}")

    # Gaps
    if gaps:
        total_gap = sum(g.get("duration_years", 0) for g in gaps)
        if total_gap > 1:
            concerns.append(f"Employment gaps detected totalling ~{total_gap:.0f} year(s)")

    if edu_gaps:
        concerns.append(f"{len(edu_gaps)} educational gap(s) detected")

    # Ensure at least one entry in each list
    if not strengths:
        strengths = ["Candidate profile submitted for review"]
    if not concerns:
        concerns  = ["No major concerns identified at this stage"]

    # Recommendation
    if total_score >= 80:
        recommendation = "Strong Hire"
    elif total_score >= 65:
        recommendation = "Hire"
    elif total_score >= 50:
        recommendation = "Maybe"
    else:
        recommendation = "No Hire"

    # Assessment paragraph
    assessment = (
        f"{name} holds a {highest_degree} and has accumulated {exp_years:.0f} year(s) of "
        f"professional experience with a {progression.lower()} career trajectory. "
        f"The candidate has {total_pubs} publication(s) and a research focus that is "
        f"{focus.lower()}. "
        f"Overall, the candidate scored {total_score:.1f}/100 placing them in the "
        f"'{tier}' tier. "
        f"Key component scores — Education: {edu_score:.0f}, Research: {res_score:.0f}, "
        f"Experience: {exp_score:.0f}, Skills: {skill_score:.0f}, "
        f"Collaboration: {collab_score:.0f}."
    )

    return {
        "assessment":      assessment,
        "strengths":       strengths,
        "concerns":        concerns,
        "recommendation":  recommendation
    }

def detect_missing(candidate_data):
    missing = []
    
    personal = candidate_data.get("personal_info", {})
    if not personal.get("email"):
        missing.append("Email address")
    if not personal.get("phone"):
        missing.append("Phone number")
    
    education = candidate_data.get("education", [])
    if not education:
        missing.append("Education history")
    else:
        for edu in education:
            if not edu.get("grade_cgpa_percentage"):
                missing.append(f"Grade/CGPA for {edu.get('degree', 'degree')}")
    
    experience = candidate_data.get("experience", [])
    if not experience:
        missing.append("Professional experience")
    
    publications = candidate_data.get("publications", [])
    if not publications:
        missing.append("Publications (if applicable)")
    
    references = candidate_data.get("references", [])
    if not references or len(references) < 2:
        missing.append("Professional references (at least 2)")
    
    return missing

def draft_email(name, post_applied, missing_items):
    if not missing_items:
        return ""
    
    items_list = "\n".join(f"• {item}" for item in missing_items)
    
    email = f"""Subject: Additional Information Required - Application for {post_applied or 'Position'}

Dear {name or 'Applicant'},

Thank you for your application for the position of {post_applied or 'the advertised role'}.

While reviewing your CV, we noticed that some information appears to be missing or incomplete. To proceed with your application, we kindly request you to provide the following details:

{items_list}

Please send us an updated CV with the above information at your earliest convenience.

We look forward to completing your application review.

Best regards,
HR Department
TALASH Recruitment System"""
    
    return email

def split_candidates(text):
    return [text]

def write_excel(all_data, xlsx_path):
    if not EXCEL_OK:
        return
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, all_data)
    
    ws_rankings = wb.create_sheet("Rankings")
    write_rankings_sheet(ws_rankings, all_data)
    
    for file_path, candidates in all_data.items():
        for idx, candidate in enumerate(candidates, 1):
            name = (candidate.get("personal_info") or {}).get("name", f"Candidate {idx}")
            safe_name = re.sub(r'[^\w\s-]', '', name)[:25]
            sheet_name = f"{safe_name}"
            
            ws = wb.create_sheet(sheet_name)
            write_candidate_sheet(ws, candidate)
    
    wb.save(xlsx_path)

def write_summary_sheet(ws, all_data):
    ws.append(["TALASH - Candidate Summary Report"])
    ws.append([])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    
    all_candidates = []
    for candidates in all_data.values():
        all_candidates.extend(candidates)
    
    ws.append(["Total Candidates", len(all_candidates)])
    ws.append(["With PhD", sum(1 for c in all_candidates if any(e.get("level") == "PhD" for e in c.get("education", [])))])
    ws.append(["With Publications", sum(1 for c in all_candidates if c.get("publications"))])
    ws.append([])
    
    ws.append(["Name", "Highest Degree", "Publications", "Experience (years)", "Ranking Score", "Tier"])
    
    for candidate in all_candidates:
        name = (candidate.get("personal_info") or {}).get("name", "Unknown")
        edu = candidate.get("education", [])
        highest = edu[-1].get("degree", "N/A") if edu else "N/A"
        pubs = len(candidate.get("publications", []))
        exp_years = candidate.get("analyses", {}).get("experience_analysis", {}).get("total_years", 0)
        score = candidate.get("ranking", {}).get("total_score", 0)
        tier = candidate.get("ranking", {}).get("ranking_tier", "Unknown")
        
        ws.append([name, highest, pubs, exp_years, score, tier])

def write_rankings_sheet(ws, all_data):
    ws.append(["Candidate Rankings"])
    ws.append([])
    ws.append([
        "Rank", "Name", "Total Score", "Tier",
        "Education", "Research", "Experience", "Skills", "Collaboration"
    ])
    
    all_candidates = []
    for candidates in all_data.values():
        all_candidates.extend(candidates)
    
    all_candidates.sort(key=lambda c: c.get("ranking", {}).get("total_score", 0), reverse=True)
    
    for rank, candidate in enumerate(all_candidates, 1):
        name = (candidate.get("personal_info") or {}).get("name", "Unknown")
        ranking = candidate.get("ranking", {})
        
        ws.append([
            rank,
            name,
            ranking.get("total_score", 0),
            ranking.get("ranking_tier", "Unknown"),
            ranking.get("education_score", 0),
            ranking.get("research_score", 0),
            ranking.get("experience_score", 0),
            ranking.get("skills_score", 0),
            ranking.get("collaboration_score", 0)
        ])

def write_candidate_sheet(ws, candidate):
    ws.append(["CANDIDATE PROFILE"])
    ws.append([])
    
    personal = candidate.get("personal_info", {})
    ws.append(["Name", personal.get("name", "")])
    ws.append(["Email", personal.get("email", "")])
    ws.append(["Phone", personal.get("phone", "")])
    ws.append(["Post Applied", personal.get("post_applied", "")])
    ws.append([])
    
    ranking = candidate.get("ranking", {})
    ws.append(["RANKING SUMMARY"])
    ws.append(["Overall Score", ranking.get("total_score", 0)])
    ws.append(["Tier", ranking.get("ranking_tier", "Unknown")])
    ws.append([])
    
    ws.append(["EDUCATION"])
    education = candidate.get("education", [])
    for edu in education:
        ws.append([
            edu.get("level", ""),
            edu.get("degree", ""),
            edu.get("institution", ""),
            edu.get("grade_cgpa_percentage", ""),
            edu.get("passing_year", "")
        ])
    ws.append([])
    
    ws.append(["PUBLICATIONS"])
    publications = candidate.get("publications", [])
    for pub in publications:
        ws.append([
            pub.get("title", ""),
            pub.get("published_in", ""),
            pub.get("year", ""),
            pub.get("type", "")
        ])
    ws.append([])
    
    summary = candidate.get("summary", {})
    ws.append(["ASSESSMENT"])
    ws.append(["Recommendation", summary.get("recommendation", "N/A")])
    ws.append([])
    ws.append(["Strengths"])
    for strength in summary.get("strengths", []):
        ws.append(["•", strength])
    ws.append([])
    ws.append(["Concerns"])
    for concern in summary.get("concerns", []):
        ws.append(["•", concern])

# ─── Main extraction job ──────────────────────────────────────────────────────

def run_extraction_job(job_id, file_paths, api_key, username):
    client = init_groq_client(api_key)
    
    def log(msg):
        with jobs_lock:
            jobs[job_id]["log"].append(f"{datetime.now().strftime('%H:%M:%S')} {msg}")
    
    try:
        with jobs_lock:
            jobs[job_id]["status"] = "running"
        
        log(f"Starting extraction for {len(file_paths)} file(s)")
        log(f"LLM mode: {'✓ Groq' if client else '✗ Disabled'}")
        
        all_data = {}
        all_flat = []
        total = 0
        
        for fpath in file_paths:
            fname = Path(fpath).name
            ext = Path(fpath).suffix.lower()
            
            log(f"Processing {fname}...")
            
            text = extract_text_from_file(fpath, client)
            if not text.strip():
                log(f"⚠ {fname} — no text extracted")
                continue
            
            log(f"✓ Text extracted from {fname} ({len(text)} chars)")
            
            is_image = ext != ".pdf"
            if is_image:
                sections = [text]
                log(f"Image file — treating as single candidate")
            else:
                sections = split_candidates(text)
                log(f"Found {len(sections)} candidate(s) in {fname}")
            
            candidates = []
            for idx, section in enumerate(sections, 1):
                log(f"Analyzing candidate {idx}/{len(sections)}...")
                
                result = process_candidate_full(client, section)
                
                if result:
                    miss = detect_missing(result)
                    name = (result.get("personal_info") or {}).get("name", "(unnamed)")
                    
                    result["_source"] = fname
                    result["_missing"] = miss
                    result["_email"] = draft_email(
                        name,
                        (result.get("personal_info") or {}).get("post_applied", ""),
                        miss
                    ) if miss else ""
                    
                    # Save to MongoDB
                    mongo_id = save_candidate_to_db(username, result)
                    if mongo_id:
                        result["_id"] = mongo_id
                        log(f"✓ Saved to database: {name}")
                    
                    candidates.append(result)
                    all_flat.append(result)
                    total += 1
                    log(f"✓ Analyzed: {name} (Score: {result.get('ranking', {}).get('total_score', 0)})")
                else:
                    log(f"✗ Failed section {idx}")
                
                if client:
                    time.sleep(0.5)
            
            if candidates:
                all_data[fpath] = candidates
            
            with jobs_lock:
                jobs[job_id]["progress"] = total
        
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_name = f"talash_{ts}.xlsx"
        json_name = f"talash_{ts}.json"
        xlsx_path = str(OUTPUT_DIR / xlsx_name)
        json_path = str(OUTPUT_DIR / json_name)
        
        if all_data:
            write_excel(all_data, xlsx_path)
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(all_flat, f, indent=2, ensure_ascii=False)
            log(f"✓ Saved {xlsx_name} and {json_name}")
        else:
            xlsx_path = json_path = ""
            log("No data extracted")
        
        stats = {
            "total": len(all_flat),
            "phd": sum(1 for c in all_flat if any(
                e.get("level", "").lower() == "phd"
                for e in (c.get("education") or [])
            )),
            "with_pub": sum(1 for c in all_flat if c.get("publications")),
            "missing": sum(1 for c in all_flat if c.get("_missing")),
            "excellent": sum(1 for c in all_flat if c.get("ranking", {}).get("ranking_tier") == "Excellent"),
            "avg_score": sum(c.get("ranking", {}).get("total_score", 0) for c in all_flat) / len(all_flat) if all_flat else 0
        }
        
        with jobs_lock:
            jobs[job_id].update({
                "status": "done",
                "candidates": all_flat,
                "total": total,
                "xlsx": xlsx_name if xlsx_path else "",
                "json": json_name if json_path else "",
                "stats": stats,
            })
        
        log(f"✓ Complete — {total} candidate(s) analyzed")
    
    except Exception as e:
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"] = str(e)
        log(f"Fatal error: {e}")

# ─── Routes — Auth ────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET"])
def login_page():
    if "username" in session:
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip().lower()
    password = data.get("password") or ""
    
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    
    users = load_users()
    user = users.get(username)
    
    if not user or user["password"] != hash_password(password):
        return jsonify({"error": "Invalid credentials"}), 401
    
    session["username"] = username
    session["display"] = user.get("display", username)
    return jsonify({"ok": True, "display": session["display"]})

@app.route("/api/signup", methods=["POST"])
def api_signup():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip().lower()
    password = data.get("password") or ""
    display = (data.get("display") or username).strip()
    
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    
    if not re.match(r"^[a-z0-9_]{3,32}$", username):
        return jsonify({"error": "Username: 3-32 chars, letters/digits/underscore"}), 400
    
    users = load_users()
    if username in users:
        return jsonify({"error": "Username already taken"}), 409
    
    users[username] = {
        "password": hash_password(password),
        "display": display,
        "created": datetime.now().isoformat()
    }
    save_users(users)
    
    session["username"] = username
    session["display"] = display
    return jsonify({"ok": True, "display": display})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me")
def api_me():
    if "username" in session:
        return jsonify({
            "username": session["username"],
            "display": session.get("display", "")
        })
    return jsonify({"username": None})

# ─── Routes — App ─────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template("index.html")

@app.route("/api/upload", methods=["POST"])
@login_required
def upload():
    files = request.files.getlist("files")
    api_key = os.environ.get("GROQ_API_KEY", "")
    
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "No files provided"}), 400
    
    job_id = str(uuid.uuid4())[:8]
    file_paths = []
    
    for f in files:
        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        
        safe = re.sub(r"[^\w\-.]", "_", f.filename)
        dest = UPLOAD_DIR / f"{job_id}_{safe}"
        f.save(str(dest))
        file_paths.append(str(dest))
    
    if not file_paths:
        return jsonify({"error": "No valid files"}), 400
    
    with jobs_lock:
        jobs[job_id] = {
            "status": "queued",
            "progress": 0,
            "total": 0,
            "candidates": [],
            "stats": {},
            "log": [],
            "xlsx": "",
            "json": "",
            "error": ""
        }
    
    username = session.get("username", "unknown")
    
    t = threading.Thread(
        target=run_extraction_job,
        args=(job_id, file_paths, api_key, username),
        daemon=True
    )
    t.start()
    
    return jsonify({"job_id": job_id})

@app.route("/api/status/<job_id>")
@login_required
def status(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
    
    if not job:
        return jsonify({"error": "Job not found"}), 404
    
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "log": job["log"][-20:],
        "stats": job["stats"],
        "xlsx": job["xlsx"],
        "json": job["json"],
        "error": job["error"],
    })

@app.route("/api/candidates")
@login_required
def all_candidates():
    """Get all candidates for current user from MongoDB"""
    username = session.get("username", "")
    candidates = get_all_candidates(username)
    return jsonify(candidates)

@app.route("/api/candidates/<job_id>")
@login_required
def candidates(job_id):
    """Get candidates from a specific job (legacy support)"""
    with jobs_lock:
        job = jobs.get(job_id)
    
    if not job:
        return jsonify({"error": "Job not found"}), 404
    
    return jsonify(job.get("candidates", []))

@app.route("/api/candidate/<candidate_id>", methods=["DELETE"])
@login_required
def delete_candidate_route(candidate_id):
    """Delete a candidate from MongoDB"""
    username = session.get("username", "")
    success = delete_candidate(username, candidate_id)
    
    if success:
        return jsonify({"ok": True})
    else:
        return jsonify({"error": "Failed to delete candidate"}), 400

@app.route("/api/download/<filename>")
@login_required
def download(filename):
    safe = re.sub(r"[^\w\-.]", "", filename)
    path = OUTPUT_DIR / safe
    
    if not path.exists():
        return jsonify({"error": "File not found"}), 404
    
    return send_file(str(path), as_attachment=True)

@app.route("/api/mode")
def extraction_mode():
    api_key = os.environ.get("GROQ_API_KEY", "")
    client = init_groq_client(api_key)
    
    return jsonify({
        "llm_available": client is not None,
        "tesseract_available": TESSERACT_OK,
        "fitz_available": FITZ_OK,
        "sklearn_available": SKLEARN_OK,
        "mongodb_available": MONGO_OK,
        "image_support": PIL_OK or (client is not None),
        "scanned_pdf_support": FITZ_OK and (TESSERACT_OK or client is not None),
    })

def calculate_progression_score(education, gaps):
    score = 50
    
    level_bonus = {"Bachelor": 10, "Master": 20, "MPhil": 25, "MS": 25, "PhD": 30, "PostDoc": 35}
    if education:
        highest_level = education[-1].get("level", "")
        score += level_bonus.get(highest_level, 0)
    
    total_gap_years = sum(g.get("duration_years", 0) for g in gaps if not g.get("justified", False))
    score -= min(20, total_gap_years * 5)
    
    if len(education) >= 3 and not gaps:
        score += 10
    
    return max(0, min(100, score))


if __name__ == "__main__":
    print("=" * 60)
    print("  TALASH Smart HR Recruitment System")
    print("  http://localhost:5000")
    print("=" * 60)
    if MONGO_OK:
        print(f"  ✓ MongoDB: {MONGO_DB_NAME}")
    else:
        print("  ✗ MongoDB: Not connected")
    print("=" * 60)
    app.run(debug=True, port=5000, use_reloader=False)

def analyze_professional_experience(experience, education):
    if not experience:
        return {
            "total_years": 0,
            "overlaps": [],
            "gaps": [],
            "progression": "Unknown",
            "consistency_score": 0
        }
    
    parsed_exp = []
    for exp in experience:
        start = parse_year(exp.get("start_date", ""))
        end = parse_year(exp.get("end_date", "")) or datetime.now().year
        
        if start:
            parsed_exp.append({
                "post": exp.get("post", ""),
                "org": exp.get("organization", ""),
                "start": start,
                "end": end,
                "original": exp
            })
    
    parsed_exp.sort(key=lambda x: x["start"])
    
    overlaps = detect_employment_overlaps(parsed_exp)
    gaps = detect_employment_gaps(parsed_exp, education)
    total_years = calculate_total_experience_years(parsed_exp)
    progression = analyze_career_progression(parsed_exp)
    consistency_score = calculate_experience_consistency(parsed_exp, overlaps, gaps)
    
    return {
        "total_years": total_years,
        "overlaps": overlaps,
        "gaps": gaps,
        "progression": progression,
        "consistency_score": consistency_score
    }

def parse_year(date_str):
    if not date_str:
        return None
    
    patterns = [
        r'(\d{4})',
        r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, date_str)
        if match:
            return int(match.group(1) if len(match.groups()) == 1 else match.group(3))
    
    return None

def detect_employment_overlaps(parsed_exp):
    overlaps = []
    
    for i in range(len(parsed_exp)):
        for j in range(i + 1, len(parsed_exp)):
            exp1 = parsed_exp[i]
            exp2 = parsed_exp[j]
            
            if exp1["start"] <= exp2["end"] and exp2["start"] <= exp1["end"]:
                overlaps.append({
                    "job1": exp1["post"],
                    "job2": exp2["post"],
                    "period": f"{max(exp1['start'], exp2['start'])}-{min(exp1['end'], exp2['end'])}"
                })
    
    return overlaps

def detect_employment_gaps(parsed_exp, education):
    gaps = []
    
    edu_end_year = 0
    if education:
        for edu in education:
            year = parse_year(edu.get("passing_year", ""))
            if year:
                edu_end_year = max(edu_end_year, year)
    
    if parsed_exp and edu_end_year:
        first_job_year = parsed_exp[0]["start"]
        gap_years = first_job_year - edu_end_year
        
        if gap_years > 1:
            gaps.append({
                "type": "Post-education gap",
                "duration_years": gap_years,
                "period": f"{edu_end_year}-{first_job_year}"
            })
    
    for i in range(len(parsed_exp) - 1):
        current_end = parsed_exp[i]["end"]
        next_start = parsed_exp[i + 1]["start"]
        gap_years = next_start - current_end
        
        if gap_years > 1:
            gaps.append({
                "type": "Between jobs",
                "duration_years": gap_years,
                "period": f"{current_end}-{next_start}"
            })
    
    return gaps

def calculate_total_experience_years(parsed_exp):
    if not parsed_exp:
        return 0
    
    periods = [(exp["start"], exp["end"]) for exp in parsed_exp]
    periods.sort()
    
    merged = []
    for start, end in periods:
        if merged and start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
        else:
            merged.append((start, end))
    
    total = sum(end - start for start, end in merged)
    return round(total, 1)

def analyze_career_progression(parsed_exp):
    if len(parsed_exp) < 2:
        return "Insufficient data"
    
    seniority_keywords = {
        "junior": 1, "assistant": 2, "associate": 3, "senior": 4,
        "lead": 5, "principal": 6, "manager": 7, "director": 8, "vp": 9, "chief": 10
    }
    
    scores = []
    for exp in parsed_exp:
        post = exp["post"].lower()
        score = 0
        for keyword, value in seniority_keywords.items():
            if keyword in post:
                score = max(score, value)
        scores.append(score)
    
    if scores[-1] > scores[0]:
        return "Progressive"
    elif scores[-1] < scores[0]:
        return "Regressive"
    else:
        return "Lateral"

def calculate_experience_consistency(parsed_exp, overlaps, gaps):
    score = 80
    
    score -= len(overlaps) * 10
    
    total_gap_years = sum(g.get("duration_years", 0) for g in gaps)
    score -= min(30, total_gap_years * 5)
    
    if len(parsed_exp) >= 3 and not gaps:
        score += 10
    
    return max(0, min(100, score))

def analyze_skill_alignment(candidate_data, client=None):
    skills = candidate_data.get("skills", {})
    experience = candidate_data.get("experience", [])
    publications = candidate_data.get("publications", [])
    
    all_skills = []
    for category in skills.values():
        if isinstance(category, list):
            all_skills.extend(category)
    
    if not all_skills:
        return {
            "evidenced_skills": [],
            "weak_evidence": [],
            "alignment_score": 0
        }
    
    evidenced = []
    weak = []
    
    for skill in all_skills:
        evidence_count = 0
        
        for exp in experience:
            post = exp.get("post", "").lower()
            resp = exp.get("responsibilities", "").lower()
            if skill.lower() in post or skill.lower() in resp:
                evidence_count += 1
        
        for pub in publications:
            title = pub.get("title", "").lower()
            if skill.lower() in title:
                evidence_count += 1
        
        if evidence_count >= 2:
            evidenced.append(skill)
        elif evidence_count == 0:
            weak.append(skill)
    
    alignment_score = (len(evidenced) / len(all_skills) * 100) if all_skills else 0
    
    return {
        "evidenced_skills": evidenced[:10],
        "weak_evidence": weak[:10],
        "alignment_score": round(alignment_score, 1)
    }

# ─── Candidate Ranking System ─────────────────────────────────────────────────

def calculate_candidate_ranking(candidate_data, analyses):
    weights = {
        "education": 0.20,
        "research": 0.35,
        "experience": 0.25,
        "skills": 0.10,
        "collaboration": 0.10
    }
    
    edu_score = analyses.get("educational_analysis", {}).get("progression_score", 0)
    
    pubs = candidate_data.get("publications", [])
    research_score = min(100, len(pubs) * 10)
    
    exp_score = analyses.get("experience_analysis", {}).get("consistency_score", 0)
    skill_score = analyses.get("skill_alignment", {}).get("alignment_score", 0)
    collab_score = analyses.get("coauthorship_analysis", {}).get("collaboration_score", 0)
    
    total_score = (
        edu_score * weights["education"] +
        research_score * weights["research"] +
        exp_score * weights["experience"] +
        skill_score * weights["skills"] +
        collab_score * weights["collaboration"]
    )
    
    return {
        "total_score": round(total_score, 1),
        "education_score": round(edu_score, 1),
        "research_score": round(research_score, 1),
        "experience_score": round(exp_score, 1),
        "skills_score": round(skill_score, 1),
        "collaboration_score": round(collab_score, 1),
        "ranking_tier": get_ranking_tier(total_score)
    }

def get_ranking_tier(score):
    if score >= 80:
        return "Excellent"
    elif score >= 65:
        return "Very Good"
    elif score >= 50:
        return "Good"
    elif score >= 35:
        return "Fair"
    else:
        return "Needs Improvement"

# ─── Comprehensive Candidate Processing ──────────────────────────────────────

def process_candidate_full(client, cv_text):
    if not client:
        return None
    
    prompt = EXTRACTION_PROMPT.format(cv_text=cv_text[:15000])
    response = call_groq(client, prompt, max_tokens=8000)
    candidate_data = parse_json_response(response)
    
    if not candidate_data:
        return None
    
    publications = candidate_data.get("publications", [])
    education = candidate_data.get("education", [])
    experience = candidate_data.get("experience", [])
    
    journal_analyses = []
    conference_analyses = []
    
    for pub in publications:
        pub_type = pub.get("type", "").lower()
        if "journal" in pub_type:
            journal_analyses.append(analyze_journal_publication(pub, client))
        elif "conference" in pub_type:
            conference_analyses.append(analyze_conference_publication(pub, client))
    
    topic_analysis = analyze_topic_variability(publications, client)
    coauthor_analysis = analyze_coauthorship(publications)
    edu_analysis = analyze_educational_profile(education, client)
    exp_analysis = analyze_professional_experience(experience, education)
    skill_analysis = analyze_skill_alignment(candidate_data, client)
    
    analyses = {
        "journal_analyses": journal_analyses,
        "conference_analyses": conference_analyses,
        "topic_variability": topic_analysis,
        "coauthorship_analysis": coauthor_analysis,
        "educational_analysis": edu_analysis,
        "experience_analysis": exp_analysis,
        "skill_alignment": skill_analysis
    }
    
    ranking = calculate_candidate_ranking(candidate_data, analyses)
    summary = generate_candidate_summary(candidate_data, analyses, ranking, client)
    
    result = {
        **candidate_data,
        "analyses": analyses,
        "ranking": ranking,
        "summary": summary
    }
    
    return result

def generate_candidate_summary(candidate_data, analyses, ranking, client):
    """Generate candidate summary — uses Groq if available, otherwise builds from data."""

    name            = (candidate_data.get("personal_info") or {}).get("name", "the candidate")
    highest_degree  = analyses.get("educational_analysis", {}).get("highest_degree", "Unknown")
    total_pubs      = len(candidate_data.get("publications", []))
    exp_years       = analyses.get("experience_analysis", {}).get("total_years", 0)
    focus           = analyses.get("topic_variability", {}).get("focus", "Unknown")
    collab_score    = analyses.get("coauthorship_analysis", {}).get("collaboration_score", 0)
    edu_score       = ranking.get("education_score", 0)
    res_score       = ranking.get("research_score", 0)
    exp_score       = ranking.get("experience_score", 0)
    skill_score     = ranking.get("skills_score", 0)
    total_score     = ranking.get("total_score", 0)
    tier            = ranking.get("ranking_tier", "Unknown")
    progression     = analyses.get("experience_analysis", {}).get("progression", "Unknown")
    evidenced_skills = analyses.get("skill_alignment", {}).get("evidenced_skills", [])
    weak_skills      = analyses.get("skill_alignment", {}).get("weak_evidence", [])
    gaps             = analyses.get("experience_analysis", {}).get("gaps", [])
    edu_gaps         = analyses.get("educational_analysis", {}).get("gaps", [])

    # ── Try Groq first ──────────────────────────────────────────────────────
    if client:
        prompt = f"""
Analyze this academic/professional candidate and generate a structured assessment.

Name: {name}
Highest Degree: {highest_degree}
Total Publications: {total_pubs}
Experience (years): {exp_years}
Research Focus: {focus}
Ranking Score: {total_score} ({tier})

Scores — Education: {edu_score}, Research: {res_score}, Experience: {exp_score}, Skills: {skill_score}, Collaboration: {collab_score}

Return ONLY valid JSON (no markdown):
{{
  "assessment": "2-3 paragraph overall assessment",
  "strengths": ["strength 1", "strength 2", "strength 3"],
  "concerns": ["concern 1", "concern 2", "concern 3"],
  "recommendation": "Strong Hire | Hire | Maybe | No Hire"
}}
"""
        response = call_groq(client, prompt, max_tokens=1500)
        result   = parse_json_response(response)

        if result and result.get("assessment") and result.get("strengths"):
            return result

    # ── Fallback: build summary from data ───────────────────────────────────
    strengths = []
    concerns  = []

    # Education
    if edu_score >= 70:
        strengths.append(f"Strong educational background — holds a {highest_degree}")
    elif edu_score >= 40:
        strengths.append(f"Solid educational foundation with a {highest_degree}")
    else:
        concerns.append(f"Educational profile needs strengthening (score: {edu_score:.0f}/100)")

    # Research
    if total_pubs > 0:
        if res_score >= 70:
            strengths.append(f"Strong research output with {total_pubs} publication(s)")
        else:
            strengths.append(f"Has research experience with {total_pubs} publication(s)")
    else:
        concerns.append("No publications found in the CV")

    # Research focus
    if focus and focus not in ("Unknown", "Insufficient data", "Limited publications"):
        strengths.append(f"Research focus is {focus.lower()}")

    # Experience
    if exp_years > 0:
        if exp_score >= 70:
            strengths.append(f"Extensive professional experience of {exp_years:.0f} year(s) with {progression.lower()} progression")
        else:
            strengths.append(f"Has {exp_years:.0f} year(s) of professional experience")
    else:
        concerns.append("Limited or no professional experience detected")

    # Collaboration
    if collab_score >= 60:
        strengths.append(f"Good collaboration record (score: {collab_score:.0f}/100)")
    elif collab_score < 30 and total_pubs > 0:
        concerns.append("Limited collaboration with other researchers")

    # Skills
    if evidenced_skills:
        strengths.append(f"Verified skills: {', '.join(evidenced_skills[:4])}")
    if weak_skills:
        concerns.append(f"Skills lacking evidence: {', '.join(weak_skills[:3])}")

    # Gaps
    if gaps:
        total_gap = sum(g.get("duration_years", 0) for g in gaps)
        if total_gap > 1:
            concerns.append(f"Employment gaps detected totalling ~{total_gap:.0f} year(s)")

    if edu_gaps:
        concerns.append(f"{len(edu_gaps)} educational gap(s) detected")

    # Ensure at least one entry in each list
    if not strengths:
        strengths = ["Candidate profile submitted for review"]
    if not concerns:
        concerns  = ["No major concerns identified at this stage"]

    # Recommendation
    if total_score >= 80:
        recommendation = "Strong Hire"
    elif total_score >= 65:
        recommendation = "Hire"
    elif total_score >= 50:
        recommendation = "Maybe"
    else:
        recommendation = "No Hire"

    # Assessment paragraph
    assessment = (
        f"{name} holds a {highest_degree} and has accumulated {exp_years:.0f} year(s) of "
        f"professional experience with a {progression.lower()} career trajectory. "
        f"The candidate has {total_pubs} publication(s) and a research focus that is "
        f"{focus.lower()}. "
        f"Overall, the candidate scored {total_score:.1f}/100 placing them in the "
        f"'{tier}' tier. "
        f"Key component scores — Education: {edu_score:.0f}, Research: {res_score:.0f}, "
        f"Experience: {exp_score:.0f}, Skills: {skill_score:.0f}, "
        f"Collaboration: {collab_score:.0f}."
    )

    return {
        "assessment":      assessment,
        "strengths":       strengths,
        "concerns":        concerns,
        "recommendation":  recommendation
    }

def detect_missing(candidate_data):
    missing = []
    
    personal = candidate_data.get("personal_info", {})
    if not personal.get("email"):
        missing.append("Email address")
    if not personal.get("phone"):
        missing.append("Phone number")
    
    education = candidate_data.get("education", [])
    if not education:
        missing.append("Education history")
    else:
        for edu in education:
            if not edu.get("grade_cgpa_percentage"):
                missing.append(f"Grade/CGPA for {edu.get('degree', 'degree')}")
    
    experience = candidate_data.get("experience", [])
    if not experience:
        missing.append("Professional experience")
    
    publications = candidate_data.get("publications", [])
    if not publications:
        missing.append("Publications (if applicable)")
    
    references = candidate_data.get("references", [])
    if not references or len(references) < 2:
        missing.append("Professional references (at least 2)")
    
    return missing

def draft_email(name, post_applied, missing_items):
    if not missing_items:
        return ""
    
    items_list = "\n".join(f"• {item}" for item in missing_items)
    
    email = f"""Subject: Additional Information Required - Application for {post_applied or 'Position'}

Dear {name or 'Applicant'},

Thank you for your application for the position of {post_applied or 'the advertised role'}.

While reviewing your CV, we noticed that some information appears to be missing or incomplete. To proceed with your application, we kindly request you to provide the following details:

{items_list}

Please send us an updated CV with the above information at your earliest convenience.

We look forward to completing your application review.

Best regards,
HR Department
TALASH Recruitment System"""
    
    return email

def split_candidates(text):
    return [text]

def write_excel(all_data, xlsx_path):
    if not EXCEL_OK:
        return
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, all_data)
    
    ws_rankings = wb.create_sheet("Rankings")
    write_rankings_sheet(ws_rankings, all_data)
    
    for file_path, candidates in all_data.items():
        for idx, candidate in enumerate(candidates, 1):
            name = (candidate.get("personal_info") or {}).get("name", f"Candidate {idx}")
            safe_name = re.sub(r'[^\w\s-]', '', name)[:25]
            sheet_name = f"{safe_name}"
            
            ws = wb.create_sheet(sheet_name)
            write_candidate_sheet(ws, candidate)
    
    wb.save(xlsx_path)

def write_summary_sheet(ws, all_data):
    ws.append(["TALASH - Candidate Summary Report"])
    ws.append([])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    
    all_candidates = []
    for candidates in all_data.values():
        all_candidates.extend(candidates)
    
    ws.append(["Total Candidates", len(all_candidates)])
    ws.append(["With PhD", sum(1 for c in all_candidates if any(e.get("level") == "PhD" for e in c.get("education", [])))])
    ws.append(["With Publications", sum(1 for c in all_candidates if c.get("publications"))])
    ws.append([])
    
    ws.append(["Name", "Highest Degree", "Publications", "Experience (years)", "Ranking Score", "Tier"])
    
    for candidate in all_candidates:
        name = (candidate.get("personal_info") or {}).get("name", "Unknown")
        edu = candidate.get("education", [])
        highest = edu[-1].get("degree", "N/A") if edu else "N/A"
        pubs = len(candidate.get("publications", []))
        exp_years = candidate.get("analyses", {}).get("experience_analysis", {}).get("total_years", 0)
        score = candidate.get("ranking", {}).get("total_score", 0)
        tier = candidate.get("ranking", {}).get("ranking_tier", "Unknown")
        
        ws.append([name, highest, pubs, exp_years, score, tier])

def write_rankings_sheet(ws, all_data):
    ws.append(["Candidate Rankings"])
    ws.append([])
    ws.append([
        "Rank", "Name", "Total Score", "Tier",
        "Education", "Research", "Experience", "Skills", "Collaboration"
    ])
    
    all_candidates = []
    for candidates in all_data.values():
        all_candidates.extend(candidates)
    
    all_candidates.sort(key=lambda c: c.get("ranking", {}).get("total_score", 0), reverse=True)
    
    for rank, candidate in enumerate(all_candidates, 1):
        name = (candidate.get("personal_info") or {}).get("name", "Unknown")
        ranking = candidate.get("ranking", {})
        
        ws.append([
            rank,
            name,
            ranking.get("total_score", 0),
            ranking.get("ranking_tier", "Unknown"),
            ranking.get("education_score", 0),
            ranking.get("research_score", 0),
            ranking.get("experience_score", 0),
            ranking.get("skills_score", 0),
            ranking.get("collaboration_score", 0)
        ])

def write_candidate_sheet(ws, candidate):
    ws.append(["CANDIDATE PROFILE"])
    ws.append([])
    
    personal = candidate.get("personal_info", {})
    ws.append(["Name", personal.get("name", "")])
    ws.append(["Email", personal.get("email", "")])
    ws.append(["Phone", personal.get("phone", "")])
    ws.append(["Post Applied", personal.get("post_applied", "")])
    ws.append([])
    
    ranking = candidate.get("ranking", {})
    ws.append(["RANKING SUMMARY"])
    ws.append(["Overall Score", ranking.get("total_score", 0)])
    ws.append(["Tier", ranking.get("ranking_tier", "Unknown")])
    ws.append([])
    
    ws.append(["EDUCATION"])
    education = candidate.get("education", [])
    for edu in education:
        ws.append([
            edu.get("level", ""),
            edu.get("degree", ""),
            edu.get("institution", ""),
            edu.get("grade_cgpa_percentage", ""),
            edu.get("passing_year", "")
        ])
    ws.append([])
    
    ws.append(["PUBLICATIONS"])
    publications = candidate.get("publications", [])
    for pub in publications:
        ws.append([
            pub.get("title", ""),
            pub.get("published_in", ""),
            pub.get("year", ""),
            pub.get("type", "")
        ])
    ws.append([])
    
    summary = candidate.get("summary", {})
    ws.append(["ASSESSMENT"])
    ws.append(["Recommendation", summary.get("recommendation", "N/A")])
    ws.append([])
    ws.append(["Strengths"])
    for strength in summary.get("strengths", []):
        ws.append(["•", strength])
    ws.append([])
    ws.append(["Concerns"])
    for concern in summary.get("concerns", []):
        ws.append(["•", concern])

# ─── Main extraction job ──────────────────────────────────────────────────────

def run_extraction_job(job_id, file_paths, api_key, username):
    client = init_groq_client(api_key)
    
    def log(msg):
        with jobs_lock:
            jobs[job_id]["log"].append(f"{datetime.now().strftime('%H:%M:%S')} {msg}")
    
    try:
        with jobs_lock:
            jobs[job_id]["status"] = "running"
        
        log(f"Starting extraction for {len(file_paths)} file(s)")
        log(f"LLM mode: {'✓ Groq' if client else '✗ Disabled'}")
        
        all_data = {}
        all_flat = []
        total = 0
        
        for fpath in file_paths:
            fname = Path(fpath).name
            ext = Path(fpath).suffix.lower()
            
            log(f"Processing {fname}...")
            
            text = extract_text_from_file(fpath, client)
            if not text.strip():
                log(f"⚠ {fname} — no text extracted")
                continue
            
            log(f"✓ Text extracted from {fname} ({len(text)} chars)")
            
            is_image = ext != ".pdf"
            if is_image:
                sections = [text]
                log(f"Image file — treating as single candidate")
            else:
                sections = split_candidates(text)
                log(f"Found {len(sections)} candidate(s) in {fname}")
            
            candidates = []
            for idx, section in enumerate(sections, 1):
                log(f"Analyzing candidate {idx}/{len(sections)}...")
                
                result = process_candidate_full(client, section)
                
                if result:
                    miss = detect_missing(result)
                    name = (result.get("personal_info") or {}).get("name", "(unnamed)")
                    
                    result["_source"] = fname
                    result["_missing"] = miss
                    result["_email"] = draft_email(
                        name,
                        (result.get("personal_info") or {}).get("post_applied", ""),
                        miss
                    ) if miss else ""
                    
                    # Save to MongoDB
                    mongo_id = save_candidate_to_db(username, result)
                    if mongo_id:
                        result["_id"] = mongo_id
                        log(f"✓ Saved to database: {name}")
                    
                    candidates.append(result)
                    all_flat.append(result)
                    total += 1
                    log(f"✓ Analyzed: {name} (Score: {result.get('ranking', {}).get('total_score', 0)})")
                else:
                    log(f"✗ Failed section {idx}")
                
                if client:
                    time.sleep(0.5)
            
            if candidates:
                all_data[fpath] = candidates
            
            with jobs_lock:
                jobs[job_id]["progress"] = total
        
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_name = f"talash_{ts}.xlsx"
        json_name = f"talash_{ts}.json"
        xlsx_path = str(OUTPUT_DIR / xlsx_name)
        json_path = str(OUTPUT_DIR / json_name)
        
        if all_data:
            write_excel(all_data, xlsx_path)
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(all_flat, f, indent=2, ensure_ascii=False)
            log(f"✓ Saved {xlsx_name} and {json_name}")
        else:
            xlsx_path = json_path = ""
            log("No data extracted")
        
        stats = {
            "total": len(all_flat),
            "phd": sum(1 for c in all_flat if any(
                e.get("level", "").lower() == "phd"
                for e in (c.get("education") or [])
            )),
            "with_pub": sum(1 for c in all_flat if c.get("publications")),
            "missing": sum(1 for c in all_flat if c.get("_missing")),
            "excellent": sum(1 for c in all_flat if c.get("ranking", {}).get("ranking_tier") == "Excellent"),
            "avg_score": sum(c.get("ranking", {}).get("total_score", 0) for c in all_flat) / len(all_flat) if all_flat else 0
        }
        
        with jobs_lock:
            jobs[job_id].update({
                "status": "done",
                "candidates": all_flat,
                "total": total,
                "xlsx": xlsx_name if xlsx_path else "",
                "json": json_name if json_path else "",
                "stats": stats,
            })
        
        log(f"✓ Complete — {total} candidate(s) analyzed")
    
    except Exception as e:
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"] = str(e)
        log(f"Fatal error: {e}")

# ─── Routes — Auth ────────────────────────────────────────────────────────────



if __name__ == "__main__":
    print("=" * 60)
    print("  TALASH Smart HR Recruitment System")
    print("  http://localhost:5000")
    print("=" * 60)
    if MONGO_OK:
        print(f"  ✓ MongoDB: {MONGO_DB_NAME}")
    else:
        print("  ✗ MongoDB: Not connected")
    print("=" * 60)
    app.run(debug=True, port=5000, use_reloader=False)
